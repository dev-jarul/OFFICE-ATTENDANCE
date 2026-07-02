import streamlit as st
import os
import pandas as pd
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# --- KONFIGURASI DATABASE EXCEL ---
FILE_EXCEL = "database_absensi.xlsx"
SHEET_PEGAWAI = "Data_Pegawai"
SHEET_LOG = "Log_Absensi"

def inisialisasi_database_profesional():
    """
    Fungsi cerdas untuk mendeteksi jika file Excel terhapus,
    lalu otomatis membuat ulang file dengan desain tabel yang sangat rapi (Styled Excel).
    """
    if not os.path.exists(FILE_EXCEL):
        wb = openpyxl.Workbook()
        
        # Aturan Desain/Styling Excel
        font_family = "Segoe UI"
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid") # Navy Blue
        font_header = Font(name=font_family, size=11, bold=True, color="FFFFFF")
        font_title = Font(name=font_family, size=16, bold=True, color="1F4E78")
        font_italic = Font(name=font_family, size=10, italic=True, color="595959")
        
        # 1. GENERATE SHEET DATA PEGAWAI
        ws_pegawai = wb.active
        ws_pegawai.title = SHEET_PEGAWAI
        ws_pegawai.views.sheetView[0].showGridLines = True
        
        # Judul Sheet Master
        ws_pegawai["A1"] = "📋 DATA MASTER PEGAWAI"
        ws_pegawai["A1"].font = font_title
        ws_pegawai["A2"] = "Daftar pegawai aktif yang terintegrasi dengan Sistem Absensi"
        ws_pegawai["A2"].font = font_italic
        
        headers_p = ["NIP", "Nama Pegawai", "Jabatan"]
        for col_num, header in enumerate(headers_p, 1):
            cell = ws_pegawai.cell(row=4, column=col_num)
            cell.value = header
            cell.font = font_header
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws_pegawai.row_dimensions[4].height = 25
        
        # Data bawaan awal saat file pertama kali digenerate
        mock_p = [
            ["19950101", "Andi Pratama", "IT Support"],
            ["19960202", "Siti Aminah", "HRD Manager"]
        ]
        for r_idx, row_data in enumerate(mock_p, 5):
            for c_idx, val in enumerate(row_data, 1):
                cell = ws_pegawai.cell(row=r_idx, column=c_idx)
                cell.value = val
                cell.alignment = Alignment(horizontal="center" if c_idx == 1 else "left", vertical="center")
                
        # 2. GENERATE SHEET LOG ABSENSI
        ws_log = wb.create_sheet(title=SHEET_LOG)
        ws_log.views.sheetView[0].showGridLines = True
        
        # Judul Sheet Log
        ws_log["A1"] = "📊 LOG REKAPITULASI ABSENSI PEGAWAI"
        ws_log["A1"].font = font_title
        ws_log["A2"] = "Catatan otomatis jam masuk, jam pulang, dan target kerja harian"
        ws_log["A2"].font = font_italic
        
        headers_l = ["Tanggal", "NIP", "Nama Pegawai", "Jabatan", "Jam Masuk", "Jam Pulang", "Target Harian"]
        for col_num, header in enumerate(headers_l, 1):
            cell = ws_log.cell(row=4, column=col_num)
            cell.value = header
            cell.font = font_header
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws_log.row_dimensions[4].height = 25
        
        # Auto-fit lebar kolom agar teks tidak terpotong (###) di Excel
        for ws in [ws_pegawai, ws_log]:
            for col in ws.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                col_letter = get_column_letter(col[0].column)
                ws.column_dimensions[col_letter].width = max(max_len + 5, 15)
                
        wb.save(FILE_EXCEL)

# Panggil fungsi inisialisasi di awal program
inisialisasi_database_profesional()

def ambil_data_pegawai():
    """Mengambil data master pegawai dari sheet Excel"""
    inisialisasi_database_profesional()
    df = pd.read_excel(FILE_EXCEL, sheet_name=SHEET_PEGAWAI, skiprows=3) # Skip 3 baris judul block excel
    df["NIP"] = df["NIP"].astype(str)
    return df

# --- SETUP HALAMAN UTAMA STREAMLIT ---
st.set_page_config(page_title="Sistem Absensi Terintegrasi", page_icon="🏢", layout="centered")
st.title("🏢 Sistem Absensi & Manajemen Pegawai")

# Navigasi Menu berbasis Tab yang responsif untuk HP & PC
tab_masuk, tab_pulang, tab_admin = st.tabs(["📝 Modul KEHADIRAN", "🚗 Modul PULANG", "🔐 Modul ADMIN"])

# Load data pegawai terbaru untuk dropdown harian
df_master_pegawai = ambil_data_pegawai()

# ==============================================================================
# 1. MODUL KEHADIRAN (ABSEN MASUK)
# ==============================================================================
with tab_masuk:
    st.subheader("Formulir Kehadiran Masuk Kerja")
    if df_master_pegawai.empty:
        st.warning("⚠️ Belum ada data pegawai di database. Silakan masukkan data pegawai lewat modul ADMIN terlebih dahulu.")
    else:
        # Membuat opsi gabungan "NIP - Nama Pegawai" untuk mempermudah pegawai
        pilihan_pegawai_masuk = df_master_pegawai.apply(lambda r: f"{r['NIP']} - {r['Nama Pegawai']}", axis=1).tolist()
        selected_profil_masuk = st.selectbox("Pilih Profil Anda (NIP - Nama):", pilihan_pegawai_masuk, key="profil_masuk")
        
        # Potong string untuk mengambil NIP-nya saja (teks sebelum tanda " - ")
        selected_nip = selected_profil_masuk.split(" - ")[0]
        
        # Logika autofill Nama & Jabatan berdasarkan NIP pilihan
        pegawai_detail = df_master_pegawai[df_master_pegawai["NIP"] == selected_nip].iloc[0]
        nama_otomatis = pegawai_detail["Nama Pegawai"]
        jabatan_otomatis = pegawai_detail["Jabatan"]
        
        st.text_input("Nama Pegawai", value=nama_otomatis, disabled=True, key="input_nama_m")
        st.text_input("Jabatan", value=jabatan_otomatis, disabled=True, key="input_jabatan_m")
        
        tanggal_hari_ini = datetime.now().strftime("%Y-%m-%d")
        st.text_input("Tanggal Hari Ini", value=tanggal_hari_ini, disabled=True, key="input_tgl_m")
        
        # Input opsional target harian harian
        target_kerja = st.text_area("Target apa yang dikerjakan hari ini? (Opsional)", placeholder="Contoh: Menguji fitur cetak rekap laporan...")
        
        st.markdown("### Konfirmasi Kehadiran")
        col_start, col_submit = st.columns(2)
        
        if "jam_masuk_recorded" not in st.session_state:
            st.session_state.jam_masuk_recorded = ""
            
        with col_start:
            if st.button("⏱️ Klik START (Rekam Waktu)", use_container_width=True, key="btn_start_m"):
                st.session_state.jam_masuk_recorded = datetime.now().strftime("%H:%M:%S")
                st.info(f"Waktu masuk direkam: {st.session_state.jam_masuk_recorded}")
                
        with col_submit:
            if st.button("🚀 SUBMIT KEHADIRAN", type="primary", use_container_width=True, key="btn_submit_m"):
                if st.session_state.jam_masuk_recorded == "":
                    st.error("Gagal! Anda harus menekan tombol 'START' terlebih dahulu untuk merekam jam kedatangan.")
                else:
                    # Tulis baris baru ke Excel Log Absensi
                    wb = openpyxl.load_workbook(FILE_EXCEL)
                    ws_log = wb[SHEET_LOG]
                    
                    target_isi = target_kerja if target_kerja.strip() != "" else "-"
                    baris_baru = [tanggal_hari_ini, selected_nip, nama_otomatis, jabatan_otomatis, st.session_state.jam_masuk_recorded, "-", target_isi]
                    ws_log.append(baris_baru)
                    wb.save(FILE_EXCEL)
                    
                    st.success(f"Sukses! Absen masuk {nama_otomatis} berhasil disubmit pada jam {st.session_state.jam_masuk_recorded}.")
                    st.session_state.jam_masuk_recorded = "" # Reset state setelah berhasil

# ==============================================================================
# 2. MODUL PULANG (ABSEN PULANG)
# ==============================================================================
with tab_pulang:
    st.subheader("Formulir Kepulangan Kerja")
    if df_master_pegawai.empty:
        st.warning("⚠️ Belum ada data pegawai di database.")
    else:
        # Membuat opsi gabungan "NIP - Nama Pegawai" untuk mempermudah pegawai saat pulang
        pilihan_pegawai_pulang = df_master_pegawai.apply(lambda r: f"{r['NIP']} - {r['Nama Pegawai']}", axis=1).tolist()
        selected_profil_pulang = st.selectbox("Pilih Profil Anda (NIP - Nama):", pilihan_pegawai_pulang, key="profil_pulang")
        
        # Potong string untuk mengambil NIP-nya saja
        selected_nip_pulang = selected_profil_pulang.split(" - ")[0]
        
        pegawai_detail_pulang = df_master_pegawai[df_master_pegawai["NIP"] == selected_nip_pulang].iloc[0]
        nama_pulang = pegawai_detail_pulang["Nama Pegawai"]
        
        st.text_input("Nama Pegawai ", value=nama_pulang, disabled=True, key="nama_p")
        tanggal_pulang_hari_ini = datetime.now().strftime("%Y-%m-%d")
        st.text_input("Tanggal ", value=tanggal_pulang_hari_ini, disabled=True, key="tgl_p")
        
        col_btn_p1, col_btn_p2 = st.columns(2)
        
        if "jam_pulang_recorded" not in st.session_state:
            st.session_state.jam_pulang_recorded = ""
            
        with col_btn_p1:
            if st.button("⏱️ Klik PULANG (Rekam Waktu)", use_container_width=True, key="btn_p1"):
                st.session_state.jam_pulang_recorded = datetime.now().strftime("%H:%M:%S")
                st.info(f"Waktu pulang direkam: {st.session_state.jam_pulang_recorded}")
                
        with col_btn_p2:
            if st.button("🛑 SUBMIT PULANG", type="primary", use_container_width=True, key="btn_p2"):
                if st.session_state.jam_pulang_recorded == "":
                    st.error("Gagal! Anda harus menekan tombol 'PULANG' terlebih dahulu untuk merekam jam kepulangan.")
                else:
                    wb = openpyxl.load_workbook(FILE_EXCEL)
                    ws_log = wb[SHEET_LOG]
                    update_berhasil = False
                    
                    # Looping cerdas: mencari baris absen masuk milik pegawai berdasarkan Tanggal & NIP hari ini
                    for row in range(5, ws_log.max_row + 1):
                        cell_tanggal = ws_log.cell(row=row, column=1).value
                        cell_nip = str(ws_log.cell(row=row, column=2).value)
                        
                        if cell_tanggal == tanggal_pulang_hari_ini and cell_nip == selected_nip_pulang:
                            # Update kolom ke-6 (Jam Pulang) pada baris yang sama
                            ws_log.cell(row=row, column=6, value=st.session_state.jam_pulang_recorded)
                            update_berhasil = True
                            break
                    
                    if update_berhasil:
                        wb.save(FILE_EXCEL)
                        st.success(f"Sukses! Jam pulang {nama_pulang} berhasil diperbarui di database master.")
                    else:
                        # Jika lupa absen masuk, buatkan baris baru agar data kepulangan tetap terekam aman
                        ws_log.append([tanggal_pulang_hari_ini, selected_nip_pulang, nama_pulang, pegawai_detail_pulang["Jabatan"], "-", st.session_state.jam_pulang_recorded, "Langsung Pulang (Tidak Absen Masuk)"])
                        wb.save(FILE_EXCEL)
                        st.warning(f"Sistem mendeteksi Anda belum absen masuk hari ini. Rekaman baru jam pulang telah dibuat otomatis.")
                    
                    st.session_state.jam_pulang_recorded = "" # Reset state setelah berhasil

# ==============================================================================
# 3. MODUL ADMIN (MANAJEMEN PEGAWAI & DOWNLOAD DATABASE)
# ==============================================================================
with tab_admin:
    st.subheader("🔐 Panel Kontrol Manajemen Admin")
    password_admin = st.text_input("Masukkan Password Kredensial Admin:", type="password")
    
    # Validasi Password Admin Terupdate
    if password_admin == "adminku121":
        st.success("Akses Terverifikasi. Selamat datang Admin.")
        st.write("---")
        
        # FITUR UTAMA: DOWNLOAD MASTER EXCEL KE FOLDER DOWNLOAD PERANGKAT
        st.markdown("#### 📥 Unduh Master Database Excel")
        try:
            with open(FILE_EXCEL, "rb") as f:
                bytes_data = f.read()
                
            st.download_button(
                label="📁 Download File database_absensi.xlsx",
                data=bytes_data,
                file_name="database_absensi.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )
            st.caption("💡 Catatan: File otomatis masuk ke folder 'Downloads' bawaan di HP maupun di PC Anda.")
        except Exception as e:
            st.error(f"Gagal menyiapkan file unduhan: {e}")
            
        st.write("---")
        
        # FITUR TAMBAH PEGAWAI VIA WEB
        st.markdown("#### ➕ Tambah Pegawai Baru Ke Sistem")
        with st.form(key="form_tambah_pegawai", clear_on_submit=True):
            new_nip = st.text_input("NIP Pegawai Baru")
            new_nama = st.text_input("Nama Lengkap Pegawai Baru")
            new_jabatan = st.text_input("Jabatan / Divisi")
            btn_tambah = st.form_submit_button("Simpan Pegawai")
            
            if btn_tambah:
                if new_nip.strip() == "" or new_nama.strip() == "" or new_jabatan.strip() == "":
                    st.error("Semua kolom wajib diisi!")
                elif new_nip in df_master_pegawai["NIP"].tolist():
                    st.error("Gagal! Pegawai dengan NIP tersebut sudah terdaftar.")
                else:
                    wb = openpyxl.load_workbook(FILE_EXCEL)
                    ws_p = wb[SHEET_PEGAWAI]
                    ws_p.append([new_nip, new_nama, new_jabatan])
                    wb.save(FILE_EXCEL)
                    st.success(f"Pegawai {new_nama} berhasil didaftarkan!")
                    st.experimental_rerun() # Menggunakan versi lama yang kompatibel dengan Python 3.7

        st.write("---")
        
        # FITUR HAPUS PEGAWAI VIA WEB
        st.markdown("#### ❌ Hapus Pegawai dari Sistem")
        if df_master_pegawai.empty:
            st.info("Tidak ada data pegawai untuk dihapus.")
        else:
            pegawai_pilihan_hapus = st.selectbox(
                "Pilih Pegawai yang ingin dihapus:",
                df_master_pegawai.apply(lambda r: f"{r['NIP']} - {r['Nama Pegawai']}", axis=1).tolist()
            )
            nip_target_hapus = pegawai_pilihan_hapus.split(" - ")[0]
            
            if st.button("🗑️ HAPUS PERMANEN", type="primary"):
                wb = openpyxl.load_workbook(FILE_EXCEL)
                ws_p = wb[SHEET_PEGAWAI]
                
                # Menghapus baris dari baris ke-5 ke bawah (area data)
                for row in range(5, ws_p.max_row + 1):
                    if str(ws_p.cell(row=row, column=1).value) == nip_target_hapus:
                        ws_p.delete_rows(row, 1)
                        break
                        
                wb.save(FILE_EXCEL)
                st.success(f"Pegawai dengan NIP {nip_target_hapus} berhasil dihapus dari sistem.")
                st.experimental_rerun()

        st.write("---")
        
        # INTEGRASI MONITORING: LIHAT LOG DATA LANGSUNG DI WEB
        st.markdown("#### 📊 Tampilan Log Aktivitas Absensi Terbaru (Real-time)")
        try:
            df_log_terbaru = pd.read_excel(FILE_EXCEL, sheet_name=SHEET_LOG, skiprows=3)
            st.dataframe(df_log_terbaru)
        except:
            st.info("Belum ada riwayat aktivitas absensi yang tercatat harian.")
            
    elif password_admin != "":
        st.error("Password Admin salah! Hak akses ditolak.")